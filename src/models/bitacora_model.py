"""
Modelo de datos para la Bitácora Geomecánica
Maneja la lógica de lectura/escritura usando SQLite (primario) y Excel (secundario).
"""
import shutil
import pandas as pd
import os
from datetime import datetime
from utils.config import (
    ARCHIVO_BITACORA, COLUMNAS_BITACORA, COLUMNAS_ESTANDAR,
    COLUMNAS_LABORES, COLUMNAS_SOSTENIMIENTO, BACKUP_DIR
)
from models.database import DatabaseManager

def _safe_concat(df_base: "pd.DataFrame", df_nuevo: "pd.DataFrame") -> "pd.DataFrame":
    """
    Concatena df_nuevo a df_base sin FutureWarning alineando columnas y dtypes.
    Maneja correctamente columnas con todos NA y columnas ausentes en alguno de los DataFrames.
    """
    if df_base.empty:
        return df_nuevo.copy()
    if df_nuevo.empty:
        return df_base.copy()

    # Alinear columnas: agregar columnas faltantes en cada DataFrame con pd.NA
    for col in df_nuevo.columns:
        if col not in df_base.columns:
            df_base = df_base.copy()
            df_base[col] = pd.NA
    for col in df_base.columns:
        if col not in df_nuevo.columns:
            df_nuevo = df_nuevo.copy()
            df_nuevo[col] = pd.NA

    # Forzar dtype object en columnas con todos NA para evitar FutureWarning
    df_nuevo = df_nuevo.copy()
    for col in df_nuevo.columns:
        if df_nuevo[col].isna().all():
            df_nuevo[col] = df_nuevo[col].astype(object)
    df_base = df_base.copy()
    for col in df_base.columns:
        if df_base[col].isna().all():
            df_base[col] = df_base[col].astype(object)

    return pd.concat([df_base, df_nuevo], ignore_index=True)


def _convertir_valor(valor, col_dtype) -> object:
    """
    Convierte un valor al dtype de la columna para asignación segura con df.at.
    Si falla, retorna el valor como string.

    Args:
        valor: Valor a convertir.
        col_dtype: dtype de la columna destino.

    Returns:
        Valor convertido o str del valor original.
    """
    try:
        if pd.api.types.is_integer_dtype(col_dtype):
            return int(valor) if str(valor).strip() not in ('', 'nan') else pd.NA
        elif pd.api.types.is_float_dtype(col_dtype):
            return float(valor) if str(valor).strip() not in ('', 'nan') else pd.NA
        else:
            return str(valor)
    except (ValueError, TypeError):
        return str(valor)


class BitacoraModel:
    """Gestiona la lógica de datos de la bitácora.

    Usa SQLite como almacenamiento primario y mantiene el archivo Excel
    sincronizado como formato secundario de exportación.
    """

    _UNDO_MAX = 5

    def __init__(self, archivo=None, db_path=None, empresa_id=1):
        self.archivo = archivo or ARCHIVO_BITACORA
        self._undo_stack: list = []
        self.db = DatabaseManager(db_path=db_path, empresa_id=empresa_id)
        self.inicializar_excel()
        self._migrar_excel_a_sqlite()
    
    def _hacer_backup(self):
        """
        Crea una copia de respaldo del archivo Excel en data/backups/.
        Si ya existe un backup del día, lo sobreescribe.
        Respeta el flag backup_automatico de la configuración.
        """
        try:
            from utils.config_manager import cargar_config
            config = cargar_config()
            if not config.get("backup_automatico", True):
                return
        except Exception:
            pass

        try:
            if not os.path.exists(self.archivo):
                return
            fecha_hoy = datetime.now().strftime("%Y-%m-%d")
            nombre_backup = BACKUP_DIR / f"bitacora_backup_{fecha_hoy}.xlsx"
            shutil.copy2(self.archivo, nombre_backup)
        except Exception as e:
            print(f"Advertencia: no se pudo crear backup: {str(e)}")

    def inicializar_excel(self):
        """Crea el archivo Excel si no existe; añade hojas/columnas faltantes si ya existe."""
        if not os.path.exists(self.archivo):
            bitacora = pd.DataFrame(columns=COLUMNAS_BITACORA)
            estandar = pd.DataFrame(columns=["RMR_min", "RMR_max", "Tipo", "Soporte"])
            labores = pd.DataFrame(columns=COLUMNAS_LABORES)
            cols_sost = self._columnas_sostenimiento_actuales()
            sostenimiento = pd.DataFrame(columns=cols_sost)

            with pd.ExcelWriter(self.archivo) as writer:
                bitacora.to_excel(writer, sheet_name="Bitacora", index=False)
                estandar.to_excel(writer, sheet_name="Estandar_Sostenimiento", index=False)
                labores.to_excel(writer, sheet_name="Labores", index=False)
                sostenimiento.to_excel(writer, sheet_name="Sostenimiento_Diario", index=False)
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self.archivo)
                sheetnames = wb.sheetnames
                wb.close()

                # ── Sostenimiento_Diario ──────────────────────────────────────
                if "Sostenimiento_Diario" not in sheetnames:
                    cols_sost = self._columnas_sostenimiento_actuales()
                    sostenimiento = pd.DataFrame(columns=cols_sost)
                    with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl") as writer:
                        sostenimiento.to_excel(writer, sheet_name="Sostenimiento_Diario", index=False)
                else:
                    try:
                        df_sost = pd.read_excel(self.archivo, sheet_name="Sostenimiento_Diario")
                        cols_sost = self._columnas_sostenimiento_actuales()
                        modified = False
                        for col in cols_sost:
                            if col not in df_sost.columns:
                                df_sost[col] = "" if col in ("Observaciones", "Tipo_Shotcrete") else 0
                                modified = True
                        # Añadir Tipo_Shotcrete si falta (retrocompatibilidad)
                        if "Tipo_Shotcrete" not in df_sost.columns:
                            df_sost["Tipo_Shotcrete"] = ""
                            modified = True
                        if modified:
                            with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                                if_sheet_exists="replace") as writer:
                                df_sost.to_excel(writer, sheet_name="Sostenimiento_Diario", index=False)
                    except Exception:
                        pass

                # ── Labores ───────────────────────────────────────────────────
                if "Labores" in sheetnames:
                    try:
                        df_lab = pd.read_excel(self.archivo, sheet_name="Labores")
                        lab_modified = False
                        for col in ("Fase", "Clasificacion_KPI"):
                            if col not in df_lab.columns:
                                df_lab[col] = ""
                                lab_modified = True
                        if lab_modified:
                            with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                                if_sheet_exists="replace") as writer:
                                df_lab.to_excel(writer, sheet_name="Labores", index=False)
                    except Exception:
                        pass
            except Exception:
                pass

    def _migrar_excel_a_sqlite(self):
        """Migra datos del Excel existente a SQLite (solo la primera vez)."""
        try:
            if not os.path.exists(self.archivo):
                return
            # Verificar si ya hay datos en SQLite para evitar migración duplicada
            registros = self.db.obtener_bitacora()
            if registros:
                return  # Ya hay datos, no migrar
            self.db.migrar_desde_excel(str(self.archivo))
        except Exception:
            pass

    def _sincronizar_a_excel(self, sheet: str = "Bitacora"):
        """Sincroniza los datos de SQLite al archivo Excel."""
        try:
            if sheet == "Bitacora":
                registros = self.db.obtener_bitacora()
                if registros:
                    df = pd.DataFrame(registros)
                    # Mantener solo columnas del Excel original
                    cols_excel = [c for c in COLUMNAS_BITACORA if c in df.columns]
                    df = df[cols_excel]
                else:
                    df = pd.DataFrame(columns=COLUMNAS_BITACORA)
                with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                    if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name="Bitacora", index=False)
            elif sheet == "Labores":
                labores = self.db.obtener_labores_guardadas()
                rows = []
                for labor_name in labores:
                    datos = self.db.obtener_datos_labor(labor_name)
                    if datos:
                        rows.append(datos)
                df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=COLUMNAS_LABORES)
                cols_excel = [c for c in COLUMNAS_LABORES if c in df.columns]
                df = df[cols_excel]
                with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                    if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name="Labores", index=False)
            elif sheet == "Sostenimiento_Diario":
                registros = self.db.obtener_sostenimiento()
                if registros:
                    df = pd.DataFrame(registros)
                    # Eliminar columnas internas de SQLite
                    for col in ("id", "created_at"):
                        if col in df.columns:
                            df = df.drop(columns=[col])
                else:
                    cols_sost = self._columnas_sostenimiento_actuales()
                    df = pd.DataFrame(columns=cols_sost)
                with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                    if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name="Sostenimiento_Diario", index=False)
        except Exception:
            pass

    def _columnas_sostenimiento_actuales(self) -> list:
        """Retorna la lista completa de columnas de Sostenimiento_Diario (base + activas)."""
        base = ["Fecha", "Turno", "Labor"]
        activos = self._columnas_sostenimiento_activas()
        fin = ["Observaciones"]
        # Columnas base de config + activas sin duplicar
        cols = base[:]
        for col in COLUMNAS_SOSTENIMIENTO:
            if col not in base and col not in fin and col not in cols:
                cols.append(col)
        for col in activos:
            if col not in cols and col not in fin:
                cols.append(col)
        cols += fin
        return cols

    def _columnas_sostenimiento_activas(self) -> list:
        """Retorna las columnas numéricas activas de sostenimiento desde config."""
        try:
            from utils.config_manager import cargar_config
            config = cargar_config()
            activos = config.get("sostenimientos_activos", [])
            return [s["columna"] for s in activos if isinstance(s, dict) and "columna" in s]
        except Exception:
            return [
                "Shotcrete_m3", "Pernos_Helicoidales", "Splitsets",
                "Mesh_Strap", "Cable_Bolting", "Marco_Acero"
            ]
    
    def _guardar_snapshot(self, sheet: str = "Bitacora"):
        """Guarda un snapshot del DataFrame en el stack de deshacer (máx 5)."""
        try:
            df = pd.read_excel(self.archivo, sheet_name=sheet)
            if len(self._undo_stack) >= self._UNDO_MAX:
                self._undo_stack.pop(0)
            self._undo_stack.append({"sheet": sheet, "data": df.to_dict(orient="list")})
        except Exception:
            pass

    def deshacer_ultima_accion(self) -> tuple:
        """
        Restaura el último snapshot del stack de deshacer.

        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        if not self._undo_stack:
            return False, "No hay acciones para deshacer"
        snapshot = self._undo_stack.pop()
        try:
            df = pd.DataFrame(snapshot["data"])
            sheet = snapshot["sheet"]
            with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
            return True, "Última acción deshecha correctamente"
        except Exception as e:
            return False, f"Error al deshacer: {str(e)}"

    def guardar_registro(self, datos):
        """
        Guarda un nuevo registro en la bitácora.
        Verifica duplicados por (Fecha, Turno, Labor) antes de guardar.
        Usa SQLite como almacenamiento primario y sincroniza a Excel.
        
        Args:
            datos (dict): Diccionario con los datos del registro
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Bitacora")
            exito, mensaje = self.db.guardar_registro(datos)
            if exito:
                self._sincronizar_a_excel("Bitacora")
                labor = datos.get("Labor", "")
                self.db.registrar_actividad("sistema", "crear_registro",
                                            f"Nuevo registro: {datos.get('Fecha', '')} - {labor}")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al guardar: {str(e)}"

    def guardar_registro_forzado(self, datos):
        """
        Guarda un registro omitiendo la verificación de duplicados.
        Usa SQLite como almacenamiento primario y sincroniza a Excel.
        
        Args:
            datos (dict): Diccionario con los datos del registro
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Bitacora")
            exito, mensaje = self.db.guardar_registro_forzado(datos)
            if exito:
                self._sincronizar_a_excel("Bitacora")
                labor = datos.get("Labor", "")
                self.db.registrar_actividad("sistema", "crear_registro",
                                            f"Registro forzado: {datos.get('Fecha', '')} - {labor}")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al guardar: {str(e)}"
    
    def obtener_bitacora(self):
        """Obtiene todos los registros de la bitácora desde SQLite."""
        try:
            registros = self.db.obtener_bitacora()
            if registros:
                df = pd.DataFrame(registros)
                # Mantener solo columnas compatibles con el formato original
                cols_mostrar = [c for c in COLUMNAS_BITACORA if c in df.columns]
                return df[cols_mostrar]
            return pd.DataFrame(columns=COLUMNAS_BITACORA)
        except Exception as e:
            print(f"Error al leer bitácora: {str(e)}")
            return pd.DataFrame(columns=COLUMNAS_BITACORA)
    
    def obtener_labores_guardadas(self):
        """Obtiene la lista de nombres de labores guardadas desde SQLite."""
        try:
            return self.db.obtener_labores_guardadas()
        except Exception:
            return []

    def _leer_labores_df(self):
        """Lee las labores desde SQLite y retorna un DataFrame."""
        try:
            labores = self.db.obtener_labores_guardadas()
            rows = []
            for labor_name in labores:
                datos = self.db.obtener_datos_labor(labor_name)
                if datos:
                    rows.append(datos)
            return pd.DataFrame(rows) if rows else pd.DataFrame(columns=COLUMNAS_LABORES)
        except Exception:
            return pd.DataFrame(columns=COLUMNAS_LABORES)

    def agregar_labor(self, nombre_labor, gsi="", rmr="", soporte="", tipo="Temporal",
                      fase="", clasificacion_kpi=""):
        """
        Agrega una nueva labor. Usa SQLite y sincroniza a Excel.
        Returns: tuple (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            exito, mensaje = self.db.agregar_labor(
                nombre_labor, gsi=gsi, rmr=rmr, soporte=soporte,
                tipo=tipo, fase=fase, clasificacion_kpi=clasificacion_kpi
            )
            if exito:
                self._sincronizar_a_excel("Labores")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al agregar labor: {str(e)}"

    def eliminar_labor(self, nombre_labor):
        """
        Elimina una labor. Usa SQLite y sincroniza a Excel.
        Returns: tuple (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            exito, mensaje = self.db.eliminar_labor(nombre_labor)
            if exito:
                self._sincronizar_a_excel("Labores")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al eliminar labor: {str(e)}"

    def obtener_datos_labor(self, nombre_labor):
        """
        Obtiene los datos técnicos de una labor del catálogo.
        Returns: dict con GSI, RMR, Soporte, Tipo o None
        """
        return self.db.obtener_datos_labor(nombre_labor)

    def obtener_labores_unicas(self):
        """
        Mantiene compatibilidad: ahora retorna las labores guardadas
        en la hoja 'Labores' en lugar de las únicas de la bitácora.
        """
        return self.obtener_labores_guardadas()
    
    def filtrar_labores(self, texto):
        """
        Filtra labores que contengan el texto. Usa SQLite.
        """
        return self.db.filtrar_labores(texto)
    
    def obtener_ultimo_registro_labor(self, labor):
        """
        Obtiene el último registro de una labor específica desde SQLite.
        
        Args:
            labor (str): Nombre de la labor
        
        Returns:
            dict: Datos del último registro o None
        """
        return self.db.obtener_ultimo_registro_labor(labor)
    
    def obtener_estandar_sostenimiento(self, sistema="RMR"):
        """Obtiene los estándares de sostenimiento desde SQLite."""
        try:
            registros = self.db.obtener_estandar_sostenimiento(sistema)
            if registros:
                return pd.DataFrame(registros)
            from utils.config_manager import columnas_estandar
            cols = columnas_estandar(sistema)
            return pd.DataFrame(columns=cols)
        except Exception:
            return pd.DataFrame(columns=["RMR_min", "RMR_max", "Tipo", "Soporte"])
    
    def recomendar_soporte(self, valor, tipo="Temporal", sistema="RMR"):
        """
        Recomienda soporte según el valor de clasificación y el tipo de labor.
        Usa SQLite como fuente primaria.

        Args:
            valor (float): Valor de la clasificación (RMR, Q, GSI, etc.)
            tipo (str): Tipo de labor ("Temporal" o "Permanente")
            sistema (str): Sistema de clasificación ("RMR", "Q", "GSI", etc.)

        Returns:
            str: Recomendación de soporte o vacío
        """
        return self.db.recomendar_soporte(valor, tipo=tipo, sistema=sistema)
    
    def guardar_estandar_sostenimiento(self, datos, sistema="RMR"):
        """
        Guarda los estándares de sostenimiento. Usa SQLite y sincroniza a Excel.
        
        Args:
            datos (list): Lista de diccionarios con los estándares
            sistema (str): Sistema de clasificación ("RMR", "Q", "GSI", etc.)
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            exito, mensaje = self.db.guardar_estandar_sostenimiento(datos, sistema)
            if exito:
                # Sincronizar a Excel
                from utils.config_manager import nombre_hoja_estandar, columnas_estandar
                hoja = nombre_hoja_estandar(sistema)
                cols = columnas_estandar(sistema)
                registros = self.db.obtener_estandar_sostenimiento(sistema)
                if registros:
                    df = pd.DataFrame(registros)
                    # Remove id column if present
                    if "id" in df.columns:
                        df = df.drop(columns=["id"])
                else:
                    df = pd.DataFrame(columns=cols)
                with pd.ExcelWriter(self.archivo, mode="a", engine="openpyxl",
                                    if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name=hoja, index=False)
            return exito, mensaje
        except Exception as e:
            return False, f"Error al guardar estándar: {str(e)}"
    
    def buscar_registros(self, labor="", fecha_inicio=None, fecha_fin=None):
        """
        Busca registros con filtros. Usa SQLite.
        
        Args:
            labor (str): Filtro por labor
            fecha_inicio (str): Filtro fecha inicio (formato dd/mm/yyyy)
            fecha_fin (str): Filtro fecha fin (formato dd/mm/yyyy)
        
        Returns:
            DataFrame: Registros filtrados (incluye id e imagen_path)
        """
        # Columnas extra necesarias para identificar registros y mostrar imágenes
        _COLS_EXTRA = ["id", "imagen_path"]
        try:
            registros = self.db.buscar_registros(labor, fecha_inicio, fecha_fin)
            if registros:
                df = pd.DataFrame(registros)
                cols_mostrar = (
                    [c for c in _COLS_EXTRA if c in df.columns]
                    + [c for c in COLUMNAS_BITACORA if c in df.columns]
                )
                return df[cols_mostrar]
            return pd.DataFrame(columns=_COLS_EXTRA + COLUMNAS_BITACORA)
        except Exception as e:
            print(f"Error al buscar: {str(e)}")
            return pd.DataFrame(columns=_COLS_EXTRA + COLUMNAS_BITACORA)

    def editar_registro(self, indice: int, datos: dict) -> tuple:
        """
        Edita un registro de la bitácora por su índice (0-based).
        Traduce el índice a ID de SQLite, edita allí y sincroniza a Excel.
        
        Args:
            indice: Índice de la fila en el DataFrame
            datos: Diccionario con los nuevos datos (solo los campos a modificar)
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Bitacora")
            # Obtener el ID del registro por índice
            registros = self.db.obtener_bitacora()
            if indice < 0 or indice >= len(registros):
                return False, "Índice fuera de rango"
            record_id = registros[indice]["id"]
            exito, mensaje = self.db.editar_registro(record_id, datos)
            if exito:
                self._sincronizar_a_excel("Bitacora")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al editar: {str(e)}"

    def editar_registro_por_id(self, record_id: int, datos: dict) -> tuple:
        """
        Edita un registro de la bitácora por su ID de base de datos.

        Args:
            record_id: ID primario del registro en SQLite
            datos: Diccionario con los nuevos datos (solo los campos a modificar)

        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Bitacora")
            exito, mensaje = self.db.editar_registro(record_id, datos)
            if exito:
                self._sincronizar_a_excel("Bitacora")
                self.db.registrar_actividad("sistema", "editar_registro",
                                            f"Registro #{record_id} editado")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al editar: {str(e)}"

    def eliminar_registro(self, indice: int) -> tuple:
        """
        Elimina un registro de la bitácora por su índice (0-based).
        Traduce el índice a ID de SQLite, elimina allí y sincroniza a Excel.
        
        Args:
            indice: Índice de la fila en el DataFrame
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Bitacora")
            registros = self.db.obtener_bitacora()
            if indice < 0 or indice >= len(registros):
                return False, "Índice fuera de rango"
            record_id = registros[indice]["id"]
            exito, mensaje = self.db.eliminar_registro(record_id)
            if exito:
                self._sincronizar_a_excel("Bitacora")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al eliminar: {str(e)}"

    def eliminar_registro_por_id(self, record_id: int) -> tuple:
        """
        Elimina un registro de la bitácora por su ID de base de datos.

        Args:
            record_id: ID primario del registro en SQLite

        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Bitacora")
            exito, mensaje = self.db.eliminar_registro(record_id)
            if exito:
                self._sincronizar_a_excel("Bitacora")
                self.db.registrar_actividad("sistema", "eliminar_registro",
                                            f"Registro #{record_id} eliminado")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al eliminar: {str(e)}"

    def exportar_historial_excel(self, df, nombre_archivo):
        """
        Exporta un DataFrame a un archivo Excel.
        
        Args:
            df: DataFrame a exportar
            nombre_archivo: Nombre/ruta del archivo destino
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            df.to_excel(nombre_archivo, index=False)
            return True, f"Archivo exportado: {nombre_archivo}"
        except Exception as e:
            return False, f"Error al exportar: {str(e)}"

    # ── Sostenimiento Diario ─────────────────────────────────────────────────

    def guardar_sostenimiento(self, datos: dict) -> tuple:
        """
        Guarda un registro de sostenimiento diario. Usa SQLite y sincroniza a Excel.
        
        Args:
            datos: Diccionario con los campos de sostenimiento
        
        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Sostenimiento_Diario")
            exito, mensaje = self.db.guardar_sostenimiento(datos)
            if exito:
                self._sincronizar_a_excel("Sostenimiento_Diario")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al guardar sostenimiento: {str(e)}"

    def guardar_sostenimiento_forzado(self, datos: dict) -> tuple:
        """
        Guarda un registro de sostenimiento omitiendo duplicados.
        Usa SQLite y sincroniza a Excel.
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Sostenimiento_Diario")
            exito, mensaje = self.db.guardar_sostenimiento_forzado(datos)
            if exito:
                self._sincronizar_a_excel("Sostenimiento_Diario")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al guardar sostenimiento: {str(e)}"

    def obtener_sostenimiento(self, fecha=None, labor=None) -> "pd.DataFrame":
        """
        Retorna registros de sostenimiento diario desde SQLite.
        
        Args:
            fecha: Filtrar por fecha exacta (string dd/mm/yyyy)
            labor: Filtrar por nombre de labor (substring)
        
        Returns:
            DataFrame con los registros
        """
        _COLS_SOST_CON_ID = ["id"] + COLUMNAS_SOSTENIMIENTO
        try:
            registros = self.db.obtener_sostenimiento(fecha=fecha, labor=labor)
            if registros:
                df = pd.DataFrame(registros)
                # Eliminar solo la columna interna de timestamp; mantener 'id' para referencias
                if "created_at" in df.columns:
                    df = df.drop(columns=["created_at"])
                return df
            return pd.DataFrame(columns=_COLS_SOST_CON_ID)
        except Exception:
            return pd.DataFrame(columns=_COLS_SOST_CON_ID)

    def editar_sostenimiento(self, indice: int, datos: dict) -> tuple:
        """
        Edita un registro de sostenimiento por índice (0-based).
        Traduce índice a ID de SQLite.
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Sostenimiento_Diario")
            registros = self.db.obtener_sostenimiento()
            if indice < 0 or indice >= len(registros):
                return False, "Índice fuera de rango"
            record_id = registros[indice]["id"]
            exito, mensaje = self.db.editar_sostenimiento(record_id, datos)
            if exito:
                self._sincronizar_a_excel("Sostenimiento_Diario")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al editar sostenimiento: {str(e)}"

    def editar_sostenimiento_por_id(self, record_id: int, datos: dict) -> tuple:
        """
        Edita un registro de sostenimiento por su ID de base de datos.

        Args:
            record_id: ID primario del registro en SQLite
            datos: Diccionario con los nuevos datos

        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Sostenimiento_Diario")
            exito, mensaje = self.db.editar_sostenimiento(record_id, datos)
            if exito:
                self._sincronizar_a_excel("Sostenimiento_Diario")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al editar sostenimiento: {str(e)}"

    def eliminar_sostenimiento(self, indice: int) -> tuple:
        """
        Elimina un registro de sostenimiento por índice (0-based).
        Traduce índice a ID de SQLite.
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Sostenimiento_Diario")
            registros = self.db.obtener_sostenimiento()
            if indice < 0 or indice >= len(registros):
                return False, "Índice fuera de rango"
            record_id = registros[indice]["id"]
            exito, mensaje = self.db.eliminar_sostenimiento(record_id)
            if exito:
                self._sincronizar_a_excel("Sostenimiento_Diario")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al eliminar sostenimiento: {str(e)}"

    def eliminar_sostenimiento_por_id(self, record_id: int) -> tuple:
        """
        Elimina un registro de sostenimiento por su ID de base de datos.

        Args:
            record_id: ID primario del registro en SQLite

        Returns:
            tuple: (éxito: bool, mensaje: str)
        """
        try:
            self._hacer_backup()
            self._guardar_snapshot("Sostenimiento_Diario")
            exito, mensaje = self.db.eliminar_sostenimiento(record_id)
            if exito:
                self._sincronizar_a_excel("Sostenimiento_Diario")
            return exito, mensaje
        except Exception as e:
            return False, f"Error al eliminar sostenimiento: {str(e)}"

    def obtener_totales_sostenimiento(self, fecha_inicio=None, fecha_fin=None,
                                      labor=None) -> "pd.DataFrame":
        """
        Agrupa y suma los totales de cada elemento de sostenimiento por labor.
        Usa datos de SQLite (via obtener_sostenimiento) como fuente primaria.
        
        Args:
            fecha_inicio: Fecha mínima (string dd/mm/yyyy)
            fecha_fin: Fecha máxima (string dd/mm/yyyy)
            labor: Filtrar por labor específica
        
        Returns:
            DataFrame agrupado por Labor con sumas de cada elemento
        """
        try:
            df = self.obtener_sostenimiento()
            if df.empty:
                return df

            # Usar columnas activas dinámicamente
            cols_num = self._columnas_sostenimiento_activas()
            # Incluir también las columnas fijas de la hoja aunque no estén en activas
            for col in COLUMNAS_SOSTENIMIENTO:
                if col not in ("Fecha", "Turno", "Labor", "Observaciones") and col not in cols_num:
                    cols_num.append(col)

            for col in cols_num:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

            if "Fecha" in df.columns:
                df["Fecha_dt"] = pd.to_datetime(df["Fecha"], format="%d/%m/%Y", errors="coerce")
                if fecha_inicio:
                    inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
                    df = df[df["Fecha_dt"] >= inicio]
                if fecha_fin:
                    fin = datetime.strptime(fecha_fin, "%d/%m/%Y")
                    df = df[df["Fecha_dt"] <= fin]

            if labor:
                df = df[df["Labor"].astype(str).str.contains(labor, case=False, na=False)]

            cols_agrupar = [c for c in cols_num if c in df.columns]
            if "Labor" not in df.columns or not cols_agrupar:
                return pd.DataFrame()

            return df.groupby("Labor")[cols_agrupar].sum().reset_index()
        except Exception as e:
            print(f"Error al obtener totales: {str(e)}")
            return pd.DataFrame()

    def archivar_periodo(self, fecha_inicio: str, fecha_fin: str) -> tuple:
        """
        Mueve registros de un rango de fechas al archivo histórico anual y los elimina.
        Usa SQLite como fuente primaria, archiva a Excel y sincroniza.

        Args:
            fecha_inicio: Fecha inicio en formato dd/mm/yyyy
            fecha_fin: Fecha fin en formato dd/mm/yyyy

        Returns:
            tuple: (éxito: bool, mensaje: str, cantidad: int)
        """
        from utils.config import DATA_DIR
        try:
            df = self.obtener_bitacora()
            if df.empty:
                return False, "No hay registros en la bitácora", 0

            df["Fecha_dt"] = pd.to_datetime(df["Fecha"], format="%d/%m/%Y", errors="coerce")
            inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            fin = datetime.strptime(fecha_fin, "%d/%m/%Y")

            mask_arch = (df["Fecha_dt"] >= inicio) & (df["Fecha_dt"] <= fin)
            df_archivar = df[mask_arch].drop(columns=["Fecha_dt"])

            if df_archivar.empty:
                return False, "No hay registros en ese período", 0

            anio = inicio.year
            archivo_hist = DATA_DIR / f"historico_{anio}.xlsx"

            if archivo_hist.exists():
                df_hist_existente = pd.read_excel(archivo_hist, sheet_name="Bitacora")
                df_hist_final = _safe_concat(df_hist_existente, df_archivar)
            else:
                df_hist_final = df_archivar

            with pd.ExcelWriter(str(archivo_hist), engine="openpyxl") as writer:
                df_hist_final.to_excel(writer, sheet_name="Bitacora", index=False)

            # Eliminar registros archivados de SQLite
            registros_sqlite = self.db.obtener_bitacora()
            self._hacer_backup()
            for reg in registros_sqlite:
                try:
                    fecha_reg = datetime.strptime(reg["Fecha"], "%d/%m/%Y")
                    if inicio <= fecha_reg <= fin:
                        self.db.eliminar_registro(reg["id"])
                except (ValueError, KeyError):
                    continue

            self._sincronizar_a_excel("Bitacora")

            cantidad = len(df_archivar)
            return True, (f"{cantidad} registro(s) archivado(s) en '{archivo_hist.name}'"), cantidad
        except Exception as e:
            return False, f"Error al archivar: {str(e)}", 0

    # ══════════════════════════════════════════════════════════════════════
    #  REGISTRO FOTOGRÁFICO – fotos asociadas a labores
    # ══════════════════════════════════════════════════════════════════════

    def guardar_foto_labor(
        self, labor: str, imagen_path: str, descripcion: str = ""
    ) -> tuple:
        """Guarda una foto asociada a una labor.

        Args:
            labor: Nombre de la labor.
            imagen_path: Ruta al archivo de imagen.
            descripcion: Descripción opcional.

        Returns:
            (True, mensaje) o (False, mensaje de error).
        """
        return self.db.guardar_foto_labor(labor, imagen_path, descripcion)

    def obtener_fotos_labor(self, labor: str) -> list[dict]:
        """Retorna las fotos del registro fotográfico asociadas a una labor.

        Args:
            labor: Nombre de la labor.

        Returns:
            Lista de dicts con los datos de cada foto.
        """
        return self.db.obtener_fotos_labor(labor)